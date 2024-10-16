import os  # Provides functions for interacting with the operating system
import shutil  # Offers high-level operations on files and collections of files
import subprocess  # Allows spawning new processes, connecting to their input/output/error pipes, and obtaining their return codes
import time  # Provides various time-related functions
import http.server  # Implements basic HTTP server (simple request handling)
import socketserver  # Provides a framework for network servers
import webbrowser  # Allows displaying web-based documents to users
import socket  # Provides low-level networking interface
import json  # Allows parsing and creating JSON data
from openpyxl import Workbook, load_workbook  # openpyxl is used for reading/writing Excel 2010 xlsx/xlsm/xltx/xltm files
import uuid  # Provides immutable UUID objects as specified in RFC 4122
import win32com.client  # Allows access to COM objects in Windows (e.g., for interacting with Microsoft Office applications)


PORT = 9002
node_path = r"C:\Program Files\nodejs\node.exe"  

hostname = socket.gethostname()
local_ip = socket.gethostbyname(hostname)

# Path for parameter weight input handling
original_excel_file = r"D:\Careers360\New Important task\Create Your Own Ranking\Base Excel File.xlsx"
temp_folder = r"D:\Careers360\New Important task\Create Your Own Ranking\Excel Load Balancer"

# Path for login data handling
login_data_folder = r"D:\Careers360\New Important task\Login Data"
login_data_file = os.path.join(login_data_folder, "login_data.xlsx")

# Ensure the folder for login data exists
os.makedirs(login_data_folder, exist_ok=True)

# Create the Excel file for login data if it doesn't exist
if not os.path.exists(login_data_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Serial Number", "Username", "University", "City", "Mobile", "Email", "Count"]) 
    wb.save(login_data_file)

html_content = """

<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>New Ranking System</title>
    <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Poppins:wght@600&display=swap">
    <style>
        body {
            font-family: Arial, sans-serif;
            background-color: #e6f7ff;
        }
        #container {
            display: flex;
            justify-content: center;
            margin-top: 50px;
        }
        #table-section {
            flex: 0 0 70%;
            text-align: center;
        }
        table {
            width: 80%;
            margin: 0 auto;
            border-collapse: collapse;
            table-layout: auto;
        }
        table, th, td {
            border: 1px solid black;
        }
        th, td {
            padding: 10px;
            text-align: center;
        }
        th {
            background-color: #f2f2f2;
        }
        td.metric {
            text-align: left;
        }
        td.college-name {
            text-align: left;
        }
        th.college-name {
            text-align: center;
        }
        
        /* Light background colors for each section */
        .highlight-blue {
            background-color: #ccd9ff; /* Light blue for TLR rows */
        }
        .highlight-orange {
            background-color: #ffebcc; /* Light orange for RP rows */
        }
        .highlight-purple {
            background-color: #e6ccff; /* Light purple for GO rows */
        }   
        .highlight-blue-dark {
            background-color: #cceeff; /* Light blue-dark for OI rows */
        }
        .highlight-pink {
            background-color: #e0e0eb; /* Light pink for PR row */
        }

        /* Darker background colors for sum rows with bold text */
        .sum-blue {
            background-color: #99b3ff; /* Darker blue for TLR sum row */
            font-weight: bold;
        }
        .sum-orange {
            background-color: #ffd699; /* Darker orange for RP sum row */
            font-weight: bold;
        }
        .sum-purple {
            background-color: #d1b3ff; /* Darker purple for GO sum row */
            font-weight: bold;
        }
        .sum-blue-dark {
            background-color: #99ddff; /* #e6f7ff Darker blue-dark for OI sum row */
            font-weight: bold;
        }
        .sum-pink {
            background-color: #c2c2d6; /* Darker pink for PR sum row */
            font-weight: bold;
        }

        /* General total row styling */
        tr.gtotal {
            background-color: #6699ff; /* #c79fef */
        }
        
        h1 {
            text-align: center;
            font-family: 'Poppins', sans-serif;
            font-size: 3em;
            font-weight: 600;
            color: #2a2a2a;
            text-transform: uppercase;
            margin-bottom: 20px;
            letter-spacing: 1px;
        }
        h2.subheading {
            text-align: center;
            font-family: 'Poppins', sans-serif;
            font-size: 1.8em;
            font-weight: 400;
            color: #555;
            margin-top: 10px;
            margin-bottom: 40px;
            letter-spacing: 1px;
            text-transform: uppercase;
        }
        h2.subheading::after {
            content: '';
            display: block;
            width: 60px;
            height: 3px;
            background-color: #66ccff;
            margin: 10px auto;
        }
        h2 {
            text-align: center;
        }
        #scoreDisplay {
            text-align: center;
            margin-top: 20px;
        }
        button {
            background-color: #007BFF;
            color: white;
            border: none;
            cursor: pointer;
            border-radius: 5px;
            transition: background-color 0.3s ease;
            padding: 10px 20px;
            font-size: 1.2em;
            margin-top: 30px;
            margin-left: 10px;
        }
        button:hover {
            background-color: #0056b3;
        }
        button.refresh-button {
            background-color: #ff6666;
        }
        button.refresh-button:hover {
            background-color: #e64a19;
        }
        button:disabled {
            background-color: #cccccc;
            cursor: not-allowed;
        }
        #error-message {
            color: red;
            font-size: 1.2em;
            text-align: center;
            margin-top: 20px;
        }
        #result-section {
            margin-top: 50px;
            text-align: center;
            display: none;
        }
        #result-table {
            width: 70%;
            margin: 0 auto;
            border-collapse: collapse;
            table-layout: auto;
        }
        #result-table th, #result-table td {
            border: 1px solid black;
            padding: 10px;
            text-align: center;
        }
        #result-table th.college-name {
            text-align: center;  /* Center align the column name for College Name */
        }
        #result-table td.college-name {
            text-align: left;  /* Left-align the data in the College Name column */
        }
        #result-table th {
            background-color: #4CAF50;
            color: white;
        }
    </style>
</head>
<body>
    <h1>Create Your Own Ranking</h1>
    <h2 class="subheading">NIRF 2024 - Engineering Ranking</h2>

    <div id="container">
        <div id="table-section">
            <table>
                <thead>
                    <tr>
                        <th>Parameters</th>
                        <th>Original Weightage</th>
                        <th>Revised Weightage</th>
                    </tr>
                </thead>
                <tbody>
                    <!-- Teaching, Learning & Resources (TLR) -->
                    <tr class="highlight-blue">
                        <td class="metric"><strong>SS</strong> - Student Strength </td>
                        <td>6</td>
                        <td><input type="number" min="0" max="100" id="weight_SS" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>
                    <tr class="highlight-blue">
                        <td class="metric"><strong>FSR</strong> - Faculty Student Ratio</td>
                        <td>9</td>
                        <td><input type="number" min="0" max="100" id="weight_FSR" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>
                    <tr class="highlight-blue">
                        <td class="metric"><strong>FQE</strong> -Faculty with PhD and Experience</td>
                        <td>6</td>
                        <td><input type="number" min="0" max="100" id="weight_FQE" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>
                    <tr class="highlight-blue">
                        <td class="metric"><strong>FRU</strong> - Financial Resources and Utilisation</td>
                        <td>9</td>
                        <td><input type="number" min="0" max="100" id="weight_FRU" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>

                    <!-- Total for TLR -->
                    <tr class="sum-blue">
                        <td><strong>TLR</strong> - Teaching, Learning & Resources</td>
                        <td><strong>30</strong></td>
                        <td><strong id="tlr_total">0</strong></td>
                    </tr>

                    <!-- Research and Professional Practice (RP) -->
                    <tr class="highlight-orange">
                        <td class="metric"><strong>PU</strong> - Publications</td>
                        <td>10.5</td>
                        <td><input type="number" min="0" max="100" id="weight_PU" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>
                    <tr class="highlight-orange">
                        <td class="metric"><strong>QP</strong> - Quality of Publications</td>
                        <td>12</td>
                        <td><input type="number" min="0" max="100" id="weight_QP" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>
                    <tr class="highlight-orange">
                        <td class="metric"><strong>IPR</strong> - Patents Published and Granted </td>
                        <td>4.5</td>
                        <td><input type="number" min="0" max="100" id="weight_IPR" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>
                    <tr class="highlight-orange">
                        <td class="metric"><strong>FPPP</strong> - Footprint of Projects and Professional Practice </td>
                        <td>3</td>
                        <td><input type="number" min="0" max="100" id="weight_FPPP" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>

                    <!-- Total for RP -->
                    <tr class="sum-orange">
                        <td><strong>RP</strong> - Research and Professional Practice</td>
                        <td><strong>30</strong></td>
                        <td><strong id="rp_total">0</strong></td>
                    </tr>

                    <!-- Graduation Outcomes (GO) -->
                    <tr class="highlight-purple">
                        <td class="metric"><strong>GPH</strong> - Placement and Higher Studies</td>
                        <td>8</td>
                        <td><input type="number" min="0" max="100" id="weight_GPH" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>
                    <tr class="highlight-purple">
                        <td class="metric"><strong>GUE</strong> - University Examination</td>
                        <td>3</td>
                        <td><input type="number" min="0" max="100" id="weight_GUE" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>
                    <tr class="highlight-purple">
                        <td class="metric"><strong>MS</strong> - Median Salary</td>
                        <td>5</td>
                        <td><input type="number" min="0" max="100" id="weight_MS" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>
                    <tr class="highlight-purple">
                        <td class="metric"><strong>GPHD</strong> - Number of Ph.D. Students Graduated </td>
                        <td>4</td>
                        <td><input type="number" min="0" max="100" id="weight_GPHD" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>

                    <!-- Total for GO -->
                    <tr class="sum-purple">
                        <td><strong>GO</strong> - Graduation Outcomes</td>
                        <td><strong>20</strong></td>
                        <td><strong id="go_total">0</strong></td>
                    </tr>

                    <!-- Outreach and Inclusivity (OI) -->
                    <tr class="highlight-blue-dark">
                        <td class="metric"><strong>RD</strong> - Students from other States/Countries (Region Diversity) </td>
                        <td>3</td>
                        <td><input type="number" min="0" max="100" id="weight_RD" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>
                    <tr class="highlight-blue-dark">
                        <td class="metric"><strong>WD</strong> - Women Diversity </td>
                        <td>3</td>
                        <td><input type="number" min="0" max="100" id="weight_WD" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>
                    <tr class="highlight-blue-dark">
                        <td class="metric"><strong>ESCS</strong> - Economically and Socially Challenged Students </td>
                        <td>2</td>
                        <td><input type="number" min="0" max="100" id="weight_ESCS" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>
                    <tr class="highlight-blue-dark">
                        <td class="metric"><strong>PCS</strong> - Facilities for Physically Challenged Student </td>
                        <td>2</td>
                        <td><input type="number" min="0" max="100" id="weight_PCS" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>

                    <!-- Total for OI -->
                    <tr class="sum-blue-dark">
                        <td><strong>OI</strong> - Outreach and Inclusivity</td>
                        <td><strong>10</strong></td>
                        <td><strong id="oi_total">0</strong></td>
                    </tr>

                    <!-- Perception (PR) -->
                    <tr class="highlight-pink">
                        <td class="metric"><strong>PR</strong> - Peer Perception</td>
                        <td>10</td>
                        <td><input type="number" min="0" max="100" id="weight_PR" oninput="validateInput(this); calculateTotal();" /></td>
                    </tr>

                    <!-- Total for PR -->
                    <tr class="sum-pink">
                        <td><strong>PR</strong> - Perception</td>
                        <td><strong>10</strong></td>
                        <td><strong id="pr_total">0</strong></td>
                    </tr>

                    <!-- Total -->
                    <tr class="gtotal">
                        <td><b>Total</b></td>
                        <td>100</td>
                        <td id="gtotal_value">0</td>
                    </tr>
                </tbody>
            </table>
            <button id="submitButton" onclick="saveData()" disabled>Submit</button>
            <button class="refresh-button" onclick="window.location.reload();">Refresh</button>
            
            <!-- Warning message -->
            <div id="warning-message" style="color: red; font-size: 1.5em; text-align: center; font-weight: bold; display: none;">
               The grand total should be exactly 100
            </div>
            
            <div id="error-message"></div>
        </div>
    </div>
    
    <div id="result-section">
        <h2>Revised Ranking</h2>
        <table id="result-table">
            <thead>
                <tr>
                    <th class="college-name">College Name</th>  <!-- Center-aligned column name for College Name -->
                    <th>Original Score</th>
                    <th>Revised Score</th>
                    <th>Revised Rank</th>
                </tr>
            </thead>
            <tbody id="result-body">
                <!-- Result rows will be dynamically generated here -->
            </tbody>
        </table>
    </div>

    <script>
        function validateInput(input) {
            var value = parseFloat(input.value);
            if (isNaN(value) || value < 0 || value > 100) {
                alert("Please enter a valid number between 0 and 100.");
                input.value = '';  // Clear invalid input
            }
            calculateTotal();
        }

        function calculateTotal() {
            var tlr_metrics = ['SS', 'FSR', 'FQE', 'FRU'];
            var rp_metrics = ['PU', 'QP', 'IPR', 'FPPP'];
            var go_metrics = ['GPH', 'GUE', 'MS', 'GPHD'];
            var oi_metrics = ['RD', 'WD', 'ESCS', 'PCS'];
            var pr_metrics = ['PR'];

            var tlr_total = 0;
            var rp_total = 0;
            var go_total = 0;
            var oi_total = 0;
            var pr_total = 0;

            // Calculate TLR total
            tlr_metrics.forEach(function (metric) {
                var weightInput = document.getElementById("weight_" + metric);
                if (weightInput && weightInput.value) {
                    tlr_total += parseFloat(weightInput.value);
                }
            });
            document.getElementById("tlr_total").innerText = tlr_total.toFixed(2);  // Update TLR row's sum

            // Calculate RP total
            rp_metrics.forEach(function (metric) {
                var weightInput = document.getElementById("weight_" + metric);
                if (weightInput && weightInput.value) {
                    rp_total += parseFloat(weightInput.value);
                }
            });
            document.getElementById("rp_total").innerText = rp_total.toFixed(2);  // Update RP row's sum

            // Calculate GO total
            go_metrics.forEach(function (metric) {
                var weightInput = document.getElementById("weight_" + metric);
                if (weightInput && weightInput.value) {
                    go_total += parseFloat(weightInput.value);
                }
            });
            document.getElementById("go_total").innerText = go_total.toFixed(2);  // Update GO row's sum

            // Calculate OI total
            oi_metrics.forEach(function (metric) {
                var weightInput = document.getElementById("weight_" + metric);
                if (weightInput && weightInput.value) {
                    oi_total += parseFloat(weightInput.value);
                }
            });
            document.getElementById("oi_total").innerText = oi_total.toFixed(2);  // Update OI row's sum

            // Calculate PR total
            pr_metrics.forEach(function (metric) {
                var weightInput = document.getElementById("weight_" + metric);
                if (weightInput && weightInput.value) {
                    pr_total += parseFloat(weightInput.value);
                }
            });
            document.getElementById("pr_total").innerText = pr_total.toFixed(2);  // Update PR row's sum

            // Update grand total
            var grand_total = tlr_total + rp_total + go_total + oi_total + pr_total;
            document.getElementById("gtotal_value").innerText = grand_total.toFixed(2);

            // Check if grand total is exactly 100
            var submitButton = document.getElementById("submitButton");
            var warningMessage = document.getElementById("warning-message");

            if (grand_total === 100) {
                submitButton.disabled = false;
                warningMessage.style.display = "none";
            } else {
                submitButton.disabled = true;
                warningMessage.style.display = "block";
                warningMessage.innerText = "Grand total must equal 100.";  // Update the warning message
            }
        }

        function saveData() {
            var submitButton = document.getElementById("submitButton");

            // Change the submit button color to green immediately on click
            submitButton.style.backgroundColor = "green";
            submitButton.innerHTML = "Loading...";  // Optionally change the button text

            var weightageData = {};
            var metrics = ['SS', 'FSR', 'FQE', 'FRU', 'PU', 'QP', 'IPR', 'FPPP', 'GPH', 'GUE', 'MS', 'GPHD', 'RD', 'WD', 'ESCS', 'PCS', 'PR'];

            var total = 0;
            var isValid = true;
            metrics.forEach(function (metric) {
                var weightInput = document.getElementById("weight_" + metric);
                if (weightInput && weightInput.value) {
                    var value = parseFloat(weightInput.value);
                    if (value < 0 || value > 100) {
                        isValid = false;
                    }
                    weightageData[metric] = value;
                    total += value;
                }
            });

            if (!isValid) {
                alert("Please ensure all inputs are between 0 and 100.");
                // Reset the button if there is an error
                submitButton.style.backgroundColor = "#007BFF";  // Reset to original color
                submitButton.innerHTML = "Submit";
                return;
            }

            fetch('/save', {
                method: 'POST',
                headers: {
                    'Content-Type': 'application/json',
                },
                body: JSON.stringify({ weightageData: weightageData }),
            }).then(response => response.json())
            .then(data => {
                var resultBody = document.getElementById("result-body");
                resultBody.innerHTML = ''; // Clear the existing table rows

                // Adding serial number
                data.results.forEach((row, index) => {
                    var newRow = `<tr>
                        <td class="college-name">${row.collegeName}</td>
                        <td>${row.originalScore}</td>  
                        <td>${row.calculatedScore}</td>
                        <td>${index + 1}</td>  <!-- Serial number -->
                    </tr>`;
                    resultBody.innerHTML += newRow;
                });

                // Display the result section
                document.getElementById("result-section").style.display = "block";

                // Scroll to the result section
                document.getElementById("result-section").scrollIntoView({ behavior: 'smooth' });

                // Final change after successful submission (optional)
                submitButton.innerHTML = "Submitted";
            })
            .catch(error => {
                console.error('Error:', error);
                alert('Error saving data.');
                // Reset the button if there is an error
                submitButton.style.backgroundColor = "#007BFF";  // Reset to original color
                submitButton.innerHTML = "Submit";
            });
        }
    </script>
</body>
</html>



"""






Login_HTML_Content = r"""
<!DOCTYPE html>
<html lang="en">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Create Your Own Ranking</title>
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@400;600&display=swap" rel="stylesheet">
    <style>
        body {
            font-family: 'Poppins', sans-serif;
            background-color: #f0f8ff;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            height: 100vh;
            margin: 0;
        }

        .title-container {
            text-align: center;
            margin-bottom: 20px;
        }

        /* Style for the Title */
        h1 {
            font-size: 2.5em;
            font-weight: 700;
            color: #000;
            letter-spacing: 1px;
            
        }

        /* Style for the Subtitle */
        h2 {
            font-size: 1.4em;
            font-weight: 400;
            color: #666;
            letter-spacing: 1px;
            margin-bottom: 30px;
        }

        .login-container {
            background-color: #add8e6;
            padding: 30px;
            border-radius: 12px;
            box-shadow: 0 8px 20px rgba(0, 0, 0, 0.1);
            width: 300px;
            text-align: center;
            position: relative;
        }

        .input-wrapper {
            position: relative;
            margin-bottom: 15px;
        }

        .input-wrapper i {
            position: absolute;
            top: 50%;
            left: 15px;
            transform: translateY(-50%);
            color: #888;
            font-size: 1.2em;
        }

        input[type="text"],
        input[type="email"],
        input[type="tel"] {
            width: 100%;
            padding: 12px 15px 12px 40px;
            border: 1px solid #ddd;
            border-radius: 8px;
            font-size: 1em;
            box-sizing: border-box;
            transition: border-color 0.3s ease-in-out;
        }

        input[type="text"]:focus,
        input[type="email"]:focus,
        input[type="tel"]:focus {
            border-color: #007BFF;
            outline: none;
        }

        button {
            width: 100%;
            padding: 12px;
            background: linear-gradient(135deg, #007BFF, #00A7FF);
            color: white;
            border: none;
            border-radius: 8px;
            font-size: 1.1em;
            cursor: pointer;
            margin-top: 20px;
            transition: background 0.3s ease-in-out;
        }

        button:hover {
            background: linear-gradient(135deg, #0056b3, #0072cc);
        }

        button:disabled {
            background: #aaa;
            cursor: not-allowed;
        }

        .login-container input[type="text"]::placeholder,
        .login-container input[type="email"]::placeholder,
        .login-container input[type="tel"]::placeholder {
            color: #bbb;
            font-size: 0.9em;
        }

        .icon {
            font-size: 16px;
        }

        /* Modal Styling */
        .modal {
            display: none;
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0, 0, 0, 0.6);
            justify-content: center;
            align-items: center;
        }

        .modal-content {
            background-color: white;
            padding: 40px;
            border-radius: 10px;
            width: 450px;
            text-align: center;
            box-shadow: 0 10px 20px rgba(0, 0, 0, 0.2);
        }

        .modal-content h3 {
            font-size: 1.7em;
            color: #333;
            margin-bottom: 15px;
        }

        .modal-content p {
            font-size: 1.1em;
            color: #666;
            margin-bottom: 25px;
        }

        .close {
            cursor: pointer;
            font-size: 1.4em;
            font-weight: bold;
            position: absolute;
            top: 15px;
            right: 20px;
            color: #333;
            background: none;
            border: none;
        }

        .close:hover {
            color: #f00;
        }

        .modal-content input[type="checkbox"] {
            margin-right: 10px;
        }

        .modal-content button {
            padding: 12px 20px;
            background-color: #007BFF;
            border: none;
            color: white;
            border-radius: 5px;
            font-size: 1.1em;
            cursor: pointer;
        }

        .modal-content button:hover {
            background-color: #0056b3;
        }
    </style>
</head>

<body>
    <!-- Title and Subtitle Container -->
    <div class="title-container">
        <h1>Create Your Own Ranking</h1>
        <h2>NIRF 2024 - Engineering Ranking</h2>
    </div>

    <!-- Login Form Card -->
    <div class="login-container" id="loginContainer">
        <form id="loginForm">
            <div class="input-wrapper">
                <i class="icon fa fa-user"></i>
                <input type="text" id="username" placeholder="Username" required>
            </div>

            <div class="input-wrapper">
                <i class="icon fa fa-university"></i>
                <input type="text" id="university" placeholder="University" required>
            </div>

            <div class="input-wrapper">
                <i class="icon fa fa-map-marker"></i>
                <input type="text" id="city" placeholder="City" required>
            </div>

            <div class="input-wrapper">
                <i class="icon fa fa-phone"></i>
                <input type="tel" id="mobile" placeholder="Mobile" required pattern="[0-9]{10}" title="Please enter exactly 10 digits.">
            </div>

            <div class="input-wrapper">
                <i class="icon fa fa-envelope"></i>
                <input type="email" id="email" placeholder="Email" required pattern="^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.(com|in|net|org)$" title="Please enter a valid email ending with .com, .in, .net, or .org.">
            </div>

            <button type="button" id="openModalButton" disabled>Submit</button>
        </form>
    </div>

    <!-- Modal -->
    <div id="disclaimerModal" class="modal">
        <div class="modal-content">
            <button class="close">&times;</button>
            <h3>Disclaimer</h3>
            <p>By proceeding, you acknowledge that this website is intended for internal use only. We are not responsible for any errors, inaccuracies, or issues that may arise.</p>
            <input type="checkbox" id="agreeCheckbox"> I agree to the terms and conditions.
            <br><br>
            <button id="loginButton" class="enabled" disabled>Login</button>
        </div>
    </div>

    <script>
        // Enable submit button when all fields are filled
        const form = document.getElementById('loginForm');
        const inputs = form.querySelectorAll('input');
        const submitButton = document.getElementById('openModalButton');

        function validateForm() {
            let isValid = true;

            inputs.forEach(input => {
                if (!input.value.trim()) {
                    isValid = false;
                }

                // Specific validation for mobile number
                if (input.id === 'mobile' && input.value.length !== 10) {
                    isValid = false;
                }

                // Specific validation for email
                if (input.id === 'email') {
                    const emailPattern = /^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.(com|in|net|org)$/;
                    if (!emailPattern.test(input.value)) {
                        isValid = false;
                    }
                }
            });

            submitButton.disabled = !isValid;
        }

        inputs.forEach(input => {
            input.addEventListener('input', validateForm);
        });

        // Handle modal opening
        document.getElementById('openModalButton').addEventListener('click', function () {
            document.getElementById('disclaimerModal').style.display = 'flex';
        });

        // Close modal
        document.querySelector('.close').addEventListener('click', function () {
            document.getElementById('disclaimerModal').style.display = 'none';
        });

        // Enable login button once checkbox is checked
        document.getElementById('agreeCheckbox').addEventListener('change', function () {
            const loginButton = document.getElementById('loginButton');
            loginButton.disabled = !this.checked;
        });

        // Handle form submission when login is clicked
        document.getElementById('loginButton').addEventListener('click', function () {
            const username = document.getElementById('username').value;
            const university = document.getElementById('university').value;
            const city = document.getElementById('city').value;
            const mobile = document.getElementById('mobile').value;
            const email = document.getElementById('email').value;

            // Validate email with proper domains like .com, .in, etc.
            const emailPattern = /^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.(com|in|net|org)$/;
            if (!emailPattern.test(email)) {
                alert("Please enter a valid email address ending with .com, .in, .net, or .org.");
                return;
            }

            // Store login session in sessionStorage
            sessionStorage.setItem('loggedIn', 'true');

            // Send login request to server
            fetch('/login', {
                method: 'POST',
                headers: {
                    'Content-Type': 'application/json',
                },
                body: JSON.stringify({
                    username: username,
                    university: university,
                    city: city,
                    mobile: mobile,
                    email: email
                }),
            })
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        window.location.href = '/main'; // Redirect to the main page
                    } else {
                        alert('Error saving your data, please try again.');
                    }
                })
                .catch(error => {
                    console.error('Error:', error);
                    alert('Error saving your data, please try again.');
                });
        });
    </script>
</body>

</html>



"""




class Handler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        if self.path == "/login":
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(Login_HTML_Content.encode('utf-8'))
        elif self.path == "/main":
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(html_content.encode('utf-8'))
        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        if self.path == '/login':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            post_data = json.loads(post_data)

            # Get user details from the form
            username = post_data.get('username')
            university = post_data.get('university')
            city = post_data.get('city')
            mobile = post_data.get('mobile')
            email = post_data.get('email')

            # Open the Excel file and check if the user already exists
            try:
                wb = load_workbook(login_data_file)
                ws = wb.active
                user_exists = False
                for row in ws.iter_rows(min_row=2, values_only=False):
                    if (row[1].value == username and row[2].value == university and row[3].value == city and 
                        row[4].value == mobile and row[5].value == email):
                        # User exists it will increase the count
                        row[6].value += 1
                        user_exists = True
                        break

                if not user_exists:
                    
                    serial_number = ws.max_row
                    ws.append([serial_number, username, university, city, mobile, email, 1])

                # Save the workbook
                wb.save(login_data_file)
                response = {'success': True}
            except Exception as e:
                print(f"Error saving login data: {e}")
                response = {'success': False}

            # Send response back to the client
            self.send_response(200)
            self.send_header("Content-type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps(response).encode('utf-8'))

        elif self.path == '/save':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            post_data = json.loads(post_data)

            weightage_data = post_data['weightageData']

            user_excel_file = os.path.join(temp_folder, f"temp_{uuid.uuid4().hex}.xlsx")
            try:
                shutil.copy(original_excel_file, user_excel_file)  # Create a copy of the original file
                print(f"Excel file created for user: {user_excel_file}")

                wb = load_workbook(user_excel_file, data_only=False)
                sheet = wb.active

                metric_cell_map = {
                    'SS': 'D4',
                    'FSR': 'E4',
                    'FQE': 'F4',
                    'FRU': 'G4',
                    'PU': 'I4',
                    'QP': 'J4',
                    'IPR': 'K4',
                    'FPPP': 'L4',
                    'GPH': 'N4',
                    'GUE': 'O4',
                    'MS': 'P4',
                    'GPHD': 'Q4',
                    'RD': 'S4',
                    'WD': 'T4',
                    'ESCS': 'U4',
                    'PCS': 'V4',
                    'PR': 'X4'
                }

                for metric, cell in metric_cell_map.items():
                    input_value = weightage_data.get(metric, 0)
                    sheet[cell].value = input_value

                wb.save(user_excel_file)

                excel = win32com.client.Dispatch("Excel.Application")
                workbook = excel.Workbooks.Open(user_excel_file)
                excel.CalculateFull()
                workbook.Save()
                workbook.Close()
                excel.Quit()

                wb = load_workbook(user_excel_file, data_only=True)
                sheet = wb.active

                college_names = sheet['B']
                calculated_scores = sheet['Z']
                original_scores = sheet['A']

                results = []
                for name, orig_score, calc_score in zip(college_names[5:], original_scores[5:], calculated_scores[5:]):
                    if name.value and calc_score.value:
                        try:
                            results.append({
                                "collegeName": name.value,
                                "originalScore": float(orig_score.value) if orig_score.value and isinstance(orig_score.value, (int, float)) else 0,
                                "calculatedScore": float(calc_score.value)
                            })
                        except ValueError:
                            print(f"Skipping invalid score for {name.value}")

                results.sort(key=lambda x: x["calculatedScore"], reverse=True)

                for result in results:
                    result["originalScore"] = f"{result['originalScore']:.2f}"
                    result["calculatedScore"] = f"{result['calculatedScore']:.2f}"

                response = {
                    "results": results
                }

                os.remove(user_excel_file)

                self.send_response(200)
                self.send_header("Content-type", "application/json")
                self.end_headers()
                self.wfile.write(json.dumps(response).encode('utf-8'))

            except Exception as e:
                print(f"Error processing the user's request: {e}")
                self.send_response(500)
                self.end_headers()
                self.wfile.write(b"Internal Server Error: " + str(e).encode('utf-8'))


def run_server():
    try:
        print(f"Starting server on http://{local_ip}:{PORT}")
        with socketserver.TCPServer(("0.0.0.0", PORT), Handler) as httpd:
            print(f"Serving on http://{local_ip}:{PORT}")
            webbrowser.open(f"http://{local_ip}:{PORT}/login")
            httpd.serve_forever()
    except Exception as e:
        print(f"Error occurred while starting the server: {e}")

run_server()